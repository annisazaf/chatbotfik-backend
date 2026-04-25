# Simpan sebagai check_xlsx.py di root folder, lalu python check_xlsx.py
import openpyxl

files = [
    'data/kurikulum/KurikulumD3SistemInformasi.xlsx',
    'data/kurikulum/KurikulumS1Informatika.xlsx',
    'data/kurikulum/KurikulumS1SainsData.xlsx',
    'data/kurikulum/KurikulumS1SistemInformasi.xlsx',
]

for f in files:
    print(f'\n=== {f} ===')
    wb = openpyxl.load_workbook(f)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f'  Sheet: "{sheet}"')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 4: break
            print(f'    {row}')